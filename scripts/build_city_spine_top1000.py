#!/usr/bin/env python3
"""Build the IESET top-1000 city spine from a local GHSL UCDB drop."""
from __future__ import annotations

import argparse
import hashlib
import importlib.util
import io
import json
import re
import sys
import tempfile
import zipfile
from dataclasses import asdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, NamedTuple

import pandas as pd
import yaml

ROOT = Path(__file__).resolve().parents[1]


def load_fetcher_base():
    spec = importlib.util.spec_from_file_location("ieset_fetcher_base", ROOT / "data" / "fetchers" / "_base.py")
    if spec is None or spec.loader is None:
        raise ImportError("Could not load data/fetchers/_base.py")
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


_FETCHER_BASE = load_fetcher_base()
FetchResult = _FETCHER_BASE.FetchResult
utc_now = _FETCHER_BASE.utc_now
utc_stamp = _FETCHER_BASE.utc_stamp

SOURCE_URL = "https://human-settlement.emergency.copernicus.eu/ghs_ucdb_2024.php"
METHODOLOGY_URL = "https://human-settlement.emergency.copernicus.eu/documents/GHSL_UCDB_R2024.pdf"
LICENSE = "CC BY 4.0 per GHSL UCDB R2024A release metadata; verify bundled metadata on refresh"
SOURCE_DATASET = "GHSL Urban Centre Database R2024A"
SOURCE_SYSTEM = "ghsl_ucdb_r2024a"


class ColumnSpec(NamedTuple):
    ghsl_city_id: str
    city_name: str
    population_2025: str
    country_name: str | None
    country_iso3: str | None
    area_km2_2025: str | None
    longitude: str | None
    latitude: str | None


def norm_col(value: object) -> str:
    return re.sub(r"[^A-Z0-9]+", "", str(value).upper())


def sha256_path(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def rel(path: Path) -> str:
    resolved = path.resolve()
    try:
        return str(resolved.relative_to(ROOT))
    except ValueError:
        return str(resolved)


def parse_fetch_utc(value: str | None) -> datetime:
    if not value:
        return utc_now()
    text = value.strip()
    if text.endswith("Z"):
        text = text[:-1] + "+00:00"
    try:
        parsed = datetime.fromisoformat(text)
    except ValueError:
        parsed = datetime.strptime(value, "%Y-%m-%dT%H%M%SZ").replace(tzinfo=timezone.utc)
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def path_arg(value: str | Path) -> Path:
    path = Path(value)
    return path if path.is_absolute() else ROOT / path


def choose_column(
    columns: list[str],
    *,
    exact: list[str],
    token_sets: list[tuple[str, ...]],
    label: str,
    required: bool = True,
) -> str | None:
    normalised = [(column, norm_col(column)) for column in columns]
    exact_norms = {norm_col(candidate) for candidate in exact}
    for column, normalised_name in normalised:
        if normalised_name in exact_norms:
            return column

    for token_set in token_sets:
        token_norms = tuple(norm_col(token) for token in token_set)
        matches = [
            column
            for column, normalised_name in normalised
            if all(token in normalised_name for token in token_norms)
        ]
        if matches:
            return sorted(matches, key=lambda item: (len(norm_col(item)), columns.index(item)))[0]

    if required:
        raise ValueError(
            f"Could not detect required {label} column. Available columns: {', '.join(columns[:80])}"
        )
    return None


def detect_columns(frame: pd.DataFrame) -> ColumnSpec:
    columns = [str(column) for column in frame.columns]
    return ColumnSpec(
        ghsl_city_id=choose_column(
            columns,
            exact=["ID_HDC_G0", "ID_UC_G0", "UC_ID", "ID_UC", "ID"],
            token_sets=[("ID", "HDC"), ("ID", "UC"), ("UC", "ID")],
            label="GHSL city id",
        ),
        city_name=choose_column(
            columns,
            exact=["GC_UCN_MAI_2025", "GC_UCN_MAI", "UC_NM_MN", "UC_NAME", "NAME", "City"],
            token_sets=[("UCN", "MAI", "2025"), ("UC", "NAME"), ("CITY",), ("NAME",)],
            label="city name",
        ),
        population_2025=choose_column(
            columns,
            exact=["GC_POP_TOT_2025", "POP_TOT_2025", "POPULATION_2025", "POPULATION"],
            token_sets=[("POP", "TOT", "2025"), ("POP", "2025"), ("POPULATION",)],
            label="2025 population",
        ),
        country_name=choose_column(
            columns,
            exact=["GC_CNT_GAD_2025", "GC_CNT_UNN_2025", "COUNTRY", "COUNTRY_NAME"],
            token_sets=[("COUNTRY", "NAME"), ("COUNTRY",), ("CNT", "2025"), ("GAD", "2025")],
            label="country name",
            required=False,
        ),
        country_iso3=choose_column(
            columns,
            exact=["GC_ISO3_2025", "ISO3", "COUNTRY_ISO3", "CNTR_CODE"],
            token_sets=[("ISO3",), ("ISO", "2025"), ("CNTR", "CODE")],
            label="country ISO3",
            required=False,
        ),
        area_km2_2025=choose_column(
            columns,
            exact=["GC_UCA_KM2_2025", "AREA_KM2_2025", "AREA_KM2"],
            token_sets=[("UCA", "KM2", "2025"), ("AREA", "KM2", "2025"), ("AREA", "KM2")],
            label="2025 area km2",
            required=False,
        ),
        longitude=choose_column(
            columns,
            exact=["GC_X_CTR_2025", "GC_X_CTR", "LONGITUDE", "LON", "X"],
            token_sets=[("LONGITUDE",), ("LON",), ("X", "CTR")],
            label="longitude",
            required=False,
        ),
        latitude=choose_column(
            columns,
            exact=["GC_Y_CTR_2025", "GC_Y_CTR", "LATITUDE", "LAT", "Y"],
            token_sets=[("LATITUDE",), ("LAT",), ("Y", "CTR")],
            label="latitude",
            required=False,
        ),
    )


def read_table(path: Path) -> tuple[pd.DataFrame, dict[str, Any]]:
    if not path.exists():
        raise FileNotFoundError(f"missing GHSL input file: {path}")

    suffix = path.suffix.lower()
    info: dict[str, Any] = {
        "input_file": rel(path),
        "input_sha256": sha256_path(path),
        "input_format": suffix.lstrip("."),
    }
    if suffix == ".csv":
        return pd.read_csv(path), info
    if suffix == ".xlsx":
        frame, sheet_name = read_xlsx_table(path)
        info["input_sheet"] = sheet_name
        return frame, info
    if suffix == ".xls":
        return pd.read_excel(path), info
    if suffix == ".zip":
        with zipfile.ZipFile(path) as zf:
            members = [
                member
                for member in sorted(zf.namelist())
                if not member.endswith("/") and not Path(member).name.startswith(".")
            ]
            for wanted_suffix in (".csv", ".xlsx", ".xls", ".gpkg"):
                for member in members:
                    if Path(member).suffix.lower() != wanted_suffix:
                        continue
                    info["input_member"] = member
                    info["input_member_format"] = wanted_suffix.lstrip(".")
                    data = zf.read(member)
                    if wanted_suffix == ".csv":
                        return pd.read_csv(io.BytesIO(data)), info
                    if wanted_suffix == ".xlsx":
                        frame, sheet_name = read_xlsx_table(io.BytesIO(data))
                        info["input_sheet"] = sheet_name
                        return frame, info
                    if wanted_suffix == ".xls":
                        return pd.read_excel(io.BytesIO(data)), info
                    try:
                        import geopandas as gpd  # type: ignore[import-not-found]
                    except ImportError as exc:
                        raise ImportError(
                            "GeoPackage members inside ZIP input require geopandas/fiona/pyogrio. "
                            "Use a GHSL CSV/XLSX member or install geospatial readers."
                        ) from exc
                    with tempfile.TemporaryDirectory() as tmpdir:
                        gpkg_path = Path(tmpdir) / Path(member).name
                        gpkg_path.write_bytes(data)
                        geoframe = gpd.read_file(gpkg_path)
                    return pd.DataFrame(geoframe.drop(columns="geometry", errors="ignore")), info
            raise ValueError(f"no CSV, XLSX, XLS, or GPKG member found in {path}")
    if suffix == ".gpkg":
        try:
            import geopandas as gpd  # type: ignore[import-not-found]
        except ImportError as exc:
            raise ImportError(
                "GeoPackage input requires geopandas/fiona/pyogrio. "
                "Use the GHSL CSV/XLSX download or install geospatial readers."
            ) from exc
        geoframe = gpd.read_file(path)
        return pd.DataFrame(geoframe.drop(columns="geometry", errors="ignore")), info
    raise ValueError(f"unsupported GHSL input format: {path.suffix}")


def read_xlsx_table(source: str | Path | io.BytesIO) -> tuple[pd.DataFrame, str]:
    try:
        import openpyxl
    except ImportError as exc:
        raise ImportError("XLSX input requires openpyxl. Use the GHSL CSV download or install openpyxl.") from exc

    workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
    selected = None
    selected_headers: list[str] = []
    for sheet in workbook.worksheets:
        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        headers = [clean_scalar(value) or "" for value in (first_row or [])]
        try:
            detect_columns(pd.DataFrame(columns=headers))
        except ValueError:
            continue
        selected = sheet
        selected_headers = headers
        break

    if selected is None:
        raise ValueError("Could not find an XLSX sheet with GHSL city id, city name, and population columns")

    rows = []
    for row in selected.iter_rows(min_row=2, values_only=True):
        values = list(row[: len(selected_headers)])
        if any(value is not None for value in values):
            rows.append(values)
    return pd.DataFrame(rows, columns=selected_headers), selected.title


def clean_scalar(value: object) -> str | None:
    if pd.isna(value):
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return text or None


def clean_text(series: pd.Series) -> pd.Series:
    return series.map(clean_scalar)


def optional_text(frame: pd.DataFrame, column: str | None) -> pd.Series:
    if column is None:
        return pd.Series([None] * len(frame), index=frame.index)
    return clean_text(frame[column])


def optional_numeric(frame: pd.DataFrame, column: str | None) -> pd.Series:
    if column is None:
        return pd.Series([pd.NA] * len(frame), index=frame.index, dtype="Float64")
    return pd.to_numeric(frame[column], errors="coerce")


def build_frames(raw: pd.DataFrame, limit: int) -> tuple[pd.DataFrame, pd.DataFrame, ColumnSpec, dict[str, Any]]:
    if limit < 1:
        raise ValueError("--limit must be a positive integer")

    spec = detect_columns(raw)
    work = pd.DataFrame(
        {
            "ghsl_city_id": clean_text(raw[spec.ghsl_city_id]),
            "city_name": clean_text(raw[spec.city_name]),
            "country_name": optional_text(raw, spec.country_name),
            "country_iso3": optional_text(raw, spec.country_iso3),
            "population_2025": pd.to_numeric(raw[spec.population_2025], errors="coerce"),
            "area_km2_2025": optional_numeric(raw, spec.area_km2_2025),
            "longitude": optional_numeric(raw, spec.longitude),
            "latitude": optional_numeric(raw, spec.latitude),
        }
    )
    before_filter = len(work)
    work = work.dropna(subset=["ghsl_city_id", "city_name", "population_2025"]).copy()
    work = work[work["population_2025"] > 0].copy()
    if work.empty:
        raise ValueError("no GHSL rows with id, city name, and positive 2025 population")

    work["country_iso3"] = work["country_iso3"].map(
        lambda value: value.upper() if isinstance(value, str) and len(value) == 3 else value
    )
    work["ieset_city_id"] = work["ghsl_city_id"].map(lambda value: f"{SOURCE_SYSTEM}:{value}")
    work = work.sort_values(
        ["population_2025", "city_name", "ghsl_city_id"],
        ascending=[False, True, True],
        kind="mergesort",
    )
    duplicate_ids = int(work["ieset_city_id"].duplicated().sum())
    work = work.drop_duplicates(subset=["ieset_city_id"], keep="first").head(limit).copy()
    work.insert(0, "city_rank_2025", range(1, len(work) + 1))
    work["population_2025"] = work["population_2025"].round().astype("Int64")
    work["density_per_km2_2025"] = work["population_2025"].astype(float) / work["area_km2_2025"]
    work.loc[~(work["area_km2_2025"] > 0), "density_per_km2_2025"] = pd.NA

    for field, source_column in spec._asdict().items():
        work[f"source_{field}_column"] = source_column
    work["source_dataset"] = SOURCE_DATASET
    work["source_system"] = SOURCE_SYSTEM
    work["morphology_note"] = "GHSL urban centre morphology; do not treat as a legal jurisdiction without crosswalks."

    universe_columns = [
        "ieset_city_id",
        "city_rank_2025",
        "ghsl_city_id",
        "city_name",
        "country_name",
        "country_iso3",
        "population_2025",
        "area_km2_2025",
        "density_per_km2_2025",
        "longitude",
        "latitude",
        "source_dataset",
        "source_system",
        "source_ghsl_city_id_column",
        "source_city_name_column",
        "source_population_2025_column",
        "source_country_name_column",
        "source_country_iso3_column",
        "source_area_km2_2025_column",
        "source_longitude_column",
        "source_latitude_column",
        "morphology_note",
    ]
    universe = work[universe_columns].reset_index(drop=True)

    crosswalks = universe[
        ["ieset_city_id", "ghsl_city_id", "city_name", "country_name", "country_iso3"]
    ].copy()
    crosswalks = crosswalks.rename(
        columns={
            "ghsl_city_id": "source_id",
            "city_name": "source_name",
            "country_name": "source_country_name",
            "country_iso3": "source_country_iso3",
        }
    )
    crosswalks.insert(1, "source_system", SOURCE_SYSTEM)
    crosswalks["match_type"] = "canonical_native_id"
    crosswalks["match_score"] = 1.0
    crosswalks["manual_review_required"] = False
    crosswalks["source_id_column"] = spec.ghsl_city_id
    crosswalks["source_name_column"] = spec.city_name

    stats = {
        "raw_rows": int(len(raw)),
        "candidate_rows_before_filter": int(before_filter),
        "ranked_city_rows": int(len(universe)),
        "duplicate_native_ids_dropped": duplicate_ids,
        "largest_city_population_2025": int(universe["population_2025"].max()) if len(universe) else None,
        "smallest_city_population_2025": int(universe["population_2025"].min()) if len(universe) else None,
        "country_name_count": int(universe["country_name"].dropna().nunique()),
        "country_iso3_count": int(universe["country_iso3"].dropna().nunique()),
        "missing_country_name_rows": int(universe["country_name"].isna().sum()),
        "missing_country_iso3_rows": int(universe["country_iso3"].isna().sum()),
        "column_detection": spec._asdict(),
    }
    return universe, crosswalks, spec, stats


def write_bundle(frame: pd.DataFrame, base_path: Path) -> dict[str, str]:
    base_path.parent.mkdir(parents=True, exist_ok=True)
    csv_path = base_path.with_suffix(".csv")
    json_path = base_path.with_suffix(".json")
    parquet_path = base_path.with_suffix(".parquet")

    frame.to_csv(csv_path, index=False)
    json_path.write_text(frame.to_json(orient="records", indent=2, force_ascii=False) + "\n")
    frame.to_parquet(parquet_path, engine="pyarrow", index=False)

    return {
        "csv_path": rel(csv_path),
        "csv_sha256": sha256_path(csv_path),
        "json_path": rel(json_path),
        "json_sha256": sha256_path(json_path),
        "parquet_path": rel(parquet_path),
        "parquet_sha256": sha256_path(parquet_path),
    }


def manifest_entry(result: FetchResult) -> dict[str, Any]:
    payload = asdict(result)
    payload["fetch_utc"] = result.fetch_utc.isoformat()
    payload["parquet_path"] = rel(result.parquet_path)
    return payload


def make_result(
    *,
    series_id: str,
    rows: int,
    fetch_ts: datetime,
    parquet_path: Path,
    parquet_sha: str,
    units: str,
    extra: dict[str, Any],
) -> FetchResult:
    return FetchResult(
        publisher="ghsl",
        series_id=series_id,
        source_url=SOURCE_URL,
        methodology_url=METHODOLOGY_URL,
        license=LICENSE,
        fetch_utc=fetch_ts,
        rows=rows,
        frequency="cross-section",
        units=units,
        currency=None,
        start_date="2025",
        end_date="2025",
        sha256=parquet_sha,
        parquet_path=parquet_path,
        extra=extra,
    )


def write_city_manifest(results: list[FetchResult], fetch_ts: datetime, manifest_dir: Path) -> Path:
    manifest_dir.mkdir(parents=True, exist_ok=True)
    stamp = utc_stamp(fetch_ts)
    path = manifest_dir / f"fetch_run_{stamp}_city_spine.yaml"
    payload = {
        "run_utc": stamp,
        "pipeline": "city_spine_top1000",
        "entries": [manifest_entry(result) for result in results],
    }
    path.write_text(yaml.safe_dump(payload, sort_keys=False))
    return path


def build_city_spine(
    *,
    input_path: Path,
    limit: int,
    output_dir: Path,
    manifest_dir: Path,
    fetch_ts: datetime,
) -> dict[str, Any]:
    raw, input_info = read_table(input_path)
    universe, crosswalks, _spec, stats = build_frames(raw, limit=limit)

    universe_artifacts = write_bundle(universe, output_dir / "city_universe_top1000")
    crosswalk_artifacts = write_bundle(crosswalks, output_dir / "city_crosswalks")

    common_extra = {
        "input": input_info,
        "construction": (
            "Rank GHSL UCDB R2024A urban centres by 2025 total population, keep the largest "
            f"{limit}, and mint stable IESET city ids as ghsl_ucdb_r2024a:<native_id>."
        ),
        "stats": stats,
    }
    results = [
        make_result(
            series_id="city_universe_top1000",
            rows=len(universe),
            fetch_ts=fetch_ts,
            parquet_path=(output_dir / "city_universe_top1000.parquet").resolve(),
            parquet_sha=universe_artifacts["parquet_sha256"],
            units="urban centres",
            extra={**common_extra, "artifacts": universe_artifacts, "columns": list(universe.columns)},
        ),
        make_result(
            series_id="city_crosswalks",
            rows=len(crosswalks),
            fetch_ts=fetch_ts,
            parquet_path=(output_dir / "city_crosswalks.parquet").resolve(),
            parquet_sha=crosswalk_artifacts["parquet_sha256"],
            units="city-source links",
            extra={**common_extra, "artifacts": crosswalk_artifacts, "columns": list(crosswalks.columns)},
        ),
    ]
    manifest = write_city_manifest(results, fetch_ts=fetch_ts, manifest_dir=manifest_dir)
    return {
        "universe": universe,
        "crosswalks": crosswalks,
        "manifest": manifest,
        "artifacts": {
            "city_universe_top1000": universe_artifacts,
            "city_crosswalks": crosswalk_artifacts,
        },
        "stats": stats,
    }


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", required=True, help="Local GHSL UCDB CSV/XLSX/GPKG file, or ZIP containing one.")
    parser.add_argument("--limit", type=int, default=1000, help="Number of ranked cities to retain.")
    parser.add_argument("--output-dir", default="data/derived", help="Directory for city universe/crosswalk artifacts.")
    parser.add_argument("--manifest-dir", default="data/manifests", help="Directory for run manifest YAML.")
    parser.add_argument("--fetch-utc", help="Optional UTC timestamp for deterministic builds/tests.")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    fetch_ts = parse_fetch_utc(args.fetch_utc)
    result = build_city_spine(
        input_path=path_arg(args.input).resolve(),
        limit=args.limit,
        output_dir=path_arg(args.output_dir).resolve(),
        manifest_dir=path_arg(args.manifest_dir).resolve(),
        fetch_ts=fetch_ts,
    )
    stats = result["stats"]
    print(
        "OK ghsl:city_universe_top1000 "
        f"rows={stats['ranked_city_rows']} country_names={stats['country_name_count']} "
        f"country_iso3={stats['country_iso3_count']} "
        f"population_range={stats['smallest_city_population_2025']}->{stats['largest_city_population_2025']}"
    )
    for artifact_set in result["artifacts"].values():
        print(f"artifact: {artifact_set['parquet_path']} sha256={artifact_set['parquet_sha256']}")
    print(f"manifest: {rel(result['manifest'])}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
