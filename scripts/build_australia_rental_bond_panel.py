#!/usr/bin/env python3
"""Build an Australian state rental-bond median-rent panel.

Primary jurisdiction: New South Wales (NSW Fair Trading / Rental Bonds Online).
NSW publishes monthly key-free XLSX micro-data files of residential rental-bond
lodgements (one bond per row: lodgement date, postcode, dwelling type, bedrooms,
weekly rent). This builder downloads one or more monthly lodgement workbooks,
aggregates the micro-data into an analysis panel of observed bond-lodgement rents,
and crosswalks the Greater Sydney metro aggregate to its GHSL urban-centre id.

Panel grain: one row per (jurisdiction, period, geography_id, dwelling_type,
bedroom_group). Rents are labelled as observed bond-lodgement weekly rents.

Source page (key-free, CC BY 4.0):
  https://www.nsw.gov.au/housing-and-construction/rental-forms-surveys-and-data/rental-bond-data
"""
from __future__ import annotations

import argparse
import hashlib
import importlib.util
import io
import sys
from dataclasses import asdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
import requests
import yaml

ROOT = Path(__file__).resolve().parents[1]


def load_fetcher_base():
    spec = importlib.util.spec_from_file_location("ieset_fetcher_base", ROOT / "data" / "fetchers" / "_base.py")
    if spec is None or spec.loader is None:
        raise ImportError("Could not load data/fetchers/_base.py")
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


_FETCHER_BASE = load_fetcher_base()
FetchResult = _FETCHER_BASE.FetchResult
utc_now = _FETCHER_BASE.utc_now
utc_stamp = _FETCHER_BASE.utc_stamp

PUBLISHER = "nsw_fair_trading_rental_bonds_online"
SERIES_ID = "australia_rental_bond_panel"
DATA_PAGE_URL = "https://www.nsw.gov.au/housing-and-construction/rental-forms-surveys-and-data/rental-bond-data"
METHODOLOGY_URL = "https://www.nsw.gov.au/housing-and-construction/rental-forms-surveys-and-data/rental-bond-data"
LICENSE = "Creative Commons Attribution 4.0 (NSW Government); verify current terms"
SOURCE_DATASET = "NSW Fair Trading residential rental bond lodgements (monthly micro-data)"

# Default monthly lodgement workbooks discovered on the NSW data page (2026-06-29).
# Stable nsw.gov.au /sites/default/files paths; key-free direct XLSX downloads.
DEFAULT_LODGEMENT_URLS = [
    "https://www.nsw.gov.au/sites/default/files/noindex/2026-06/rentalbond_lodgements_may_2026.xlsx",
    "https://www.nsw.gov.au/sites/default/files/noindex/2026-05/rentalbond_lodgements_april_2026.xlsx",
    "https://www.nsw.gov.au/sites/default/files/noindex/2026-04/rentalbond_lodgements_march_2026.xlsx",
    "https://www.nsw.gov.au/sites/default/files/noindex/2026-03/rentalbond_lodgements_february_2026.xlsx",
    "https://www.nsw.gov.au/sites/default/files/noindex/2026-02/rentalbond_lodgements_january_2026.xlsx",
    "https://www.nsw.gov.au/sites/default/files/noindex/2026-01/rentalbond_lodgements_december_2025.xlsx",
    "https://www.nsw.gov.au/sites/default/files/noindex/2025-12/rentalbond_lodgements_november_2025.xlsx",
    "https://www.nsw.gov.au/sites/default/files/noindex/2025-11/rentalbond_lodgements_october_2025.xlsx",
    "https://www.nsw.gov.au/sites/default/files/noindex/2025-10/rentalbond_lodgements_september25.xlsx",
    "https://www.nsw.gov.au/sites/default/files/noindex/2025-09/rentalbond_lodgements_august_2025.xlsx",
    "https://www.nsw.gov.au/sites/default/files/noindex/2025-08/rental-bond-lodgement-data-july-2025.xlsx",
    "https://www.nsw.gov.au/sites/default/files/noindex/2025-08/rentalbond_lodgements_june_2025_0.xlsx",
]

HEADERS = {"User-Agent": "Mozilla/5.0 (IESET city-level data builder)"}

DWELLING_TYPE_LABELS = {
    "F": "flat_unit",
    "H": "house",
    "T": "terrace_townhouse_semi",
    "O": "other",
    "U": "unknown",
}

# ABS Greater Sydney (Greater Capital City Statistical Area) postcode coverage.
# Inclusive integer ranges. Postcodes outside these ranges (but valid NSW 2xxx)
# are treated as Rest of NSW.
GREATER_SYDNEY_POSTCODE_RANGES = [
    (2000, 2249),  # Sydney inner, eastern suburbs, north shore, northern beaches
    (2555, 2574),  # Macarthur / Camden / Campbelltown
    (2740, 2786),  # Blue Mountains, Penrith, Hawkesbury
    (2745, 2770),  # outer western Sydney (overlaps above; kept explicit)
]

# Manual crosswalk of the Greater Sydney aggregate to its GHSL urban-centre name
# in the top-1000 city spine (mirrors the Taiwan builder's manual-name pattern).
GEOGRAPHY_GHSL_CITY_NAME = {
    "greater_sydney": "Sydney",
}


def rel(path: Path) -> str:
    resolved = path.resolve()
    try:
        return str(resolved.relative_to(ROOT))
    except ValueError:
        return str(resolved)


def sha256_bytes(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def sha256_path(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def parse_fetch_utc(value: str | None) -> datetime:
    if not value:
        return utc_now()
    text = value.strip()
    if text.endswith("Z"):
        text = text[:-1] + "+00:00"
    try:
        parsed = datetime.fromisoformat(text)
    except ValueError:
        parsed = datetime.strptime(value, "%Y-%m-%dT%H%M%SZ").replace(tzinfo=timezone.utc)
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def path_arg(value: str | Path) -> Path:
    path = Path(value)
    return path if path.is_absolute() else ROOT / path


def fetch_bytes(url: str, timeout: int = 180) -> bytes:
    response = requests.get(url, timeout=timeout, headers=HEADERS)
    response.raise_for_status()
    return response.content


def parse_int(value: object) -> int | None:
    if value is None:
        return None
    text = str(value).strip().replace(",", "")
    if text in {"", "U", "u", "nan", "None", "-"}:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def parse_rent(value: object) -> float | None:
    if value is None:
        return None
    text = str(value).strip().replace(",", "").replace("$", "")
    if text in {"", "U", "u", "nan", "None", "-"}:
        return None
    try:
        rent = float(text)
    except ValueError:
        return None
    # Plausibility floor/ceiling for a weekly rent in AUD; keep observed extremes
    # within a sane band so a stray data-entry artefact cannot poison the median.
    if rent <= 0 or rent > 50000:
        return None
    return rent


def postcode_geography(postcode: int | None) -> str | None:
    if postcode is None:
        return None
    for lo, hi in GREATER_SYDNEY_POSTCODE_RANGES:
        if lo <= postcode <= hi:
            return "greater_sydney"
    if 2000 <= postcode <= 2999 or 2600 <= postcode <= 2620:
        return "rest_of_nsw"
    return "rest_of_nsw"


def bedroom_group(bedrooms: int | None) -> str:
    if bedrooms is None:
        return "unknown"
    if bedrooms <= 0:
        return "0_studio"
    if bedrooms >= 5:
        return "5_plus"
    return str(bedrooms)


def read_lodgement_workbook(payload: bytes, source_url: str, source_file: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    try:
        sheet_name = next(
            (s for s in wb.sheetnames if "lodg" in s.lower()),
            wb.sheetnames[0],
        )
        ws = wb[sheet_name]
        header_idx = None
        rows: list[dict[str, Any]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if header_idx is None:
                cells = [str(c).strip().lower() if c is not None else "" for c in row]
                if "lodgement date" in cells and "weekly rent" in cells:
                    header_idx = i
                    col = {name: cells.index(name) for name in
                           ("lodgement date", "postcode", "dwelling type", "bedrooms", "weekly rent")}
                continue
            if row is None or all(c is None for c in row):
                continue
            lodge_date = row[col["lodgement date"]]
            if lodge_date is None:
                continue
            if isinstance(lodge_date, datetime):
                period = f"{lodge_date.year:04d}-{lodge_date.month:02d}"
            else:
                ts = pd.to_datetime(str(lodge_date), errors="coerce")
                if pd.isna(ts):
                    continue
                period = f"{ts.year:04d}-{ts.month:02d}"
            postcode = parse_int(row[col["postcode"]])
            geography = postcode_geography(postcode)
            if geography is None:
                continue
            dwelling_code = row[col["dwelling type"]]
            dwelling_code = str(dwelling_code).strip().upper() if dwelling_code is not None else "U"
            if dwelling_code not in DWELLING_TYPE_LABELS:
                dwelling_code = "U"
            bedrooms = parse_int(row[col["bedrooms"]])
            rent = parse_rent(row[col["weekly rent"]])
            rows.append(
                {
                    "period": period,
                    "geography_id": geography,
                    "postcode": postcode,
                    "dwelling_type_code": dwelling_code,
                    "bedroom_group": bedroom_group(bedrooms),
                    "bedrooms_raw": bedrooms,
                    "weekly_rent_aud": rent,
                    "source_file": source_file,
                    "source_url": source_url,
                }
            )
        if header_idx is None:
            raise ValueError(f"No recognisable header row in {source_file}")
        return pd.DataFrame(rows)
    finally:
        wb.close()


def aggregate_panel(micro: pd.DataFrame) -> pd.DataFrame:
    # Build per-geography rollups: greater_sydney + rest_of_nsw (postcode-derived),
    # and a synthetic nsw_state aggregate covering all lodgements.
    frames = [micro]
    state = micro.copy()
    state["geography_id"] = "nsw_state"
    frames.append(state)
    stacked = pd.concat(frames, ignore_index=True)

    group_cols = ["jurisdiction_code", "period", "geography_id", "dwelling_type_code", "bedroom_group"]
    stacked["jurisdiction_code"] = "AU-NSW"

    records: list[dict[str, Any]] = []
    for keys, grp in stacked.groupby(["period", "geography_id", "dwelling_type_code", "bedroom_group"], sort=True):
        period, geography_id, dwelling_code, bed_group = keys
        rents = grp["weekly_rent_aud"].dropna()
        bond_lodgements = int(len(grp))
        rent_observations = int(rents.shape[0])
        records.append(
            {
                "jurisdiction_code": "AU-NSW",
                "jurisdiction_name": "New South Wales",
                "country_name": "Australia",
                "country_iso3": "AUS",
                "period": period,
                "period_type": "month",
                "geography_id": geography_id,
                "geography_label": {
                    "greater_sydney": "Greater Sydney (GCCSA, postcode-derived)",
                    "rest_of_nsw": "Rest of New South Wales (postcode-derived)",
                    "nsw_state": "New South Wales (all lodgements)",
                }[geography_id],
                "dwelling_type_code": dwelling_code,
                "dwelling_type_label": DWELLING_TYPE_LABELS[dwelling_code],
                "bedroom_group": bed_group,
                "bond_lodgement_count": bond_lodgements,
                "rent_observation_count": rent_observations,
                "median_weekly_rent_aud": round(float(rents.median()), 2) if rent_observations else None,
                "mean_weekly_rent_aud": round(float(rents.mean()), 2) if rent_observations else None,
                "p25_weekly_rent_aud": round(float(rents.quantile(0.25)), 2) if rent_observations else None,
                "p75_weekly_rent_aud": round(float(rents.quantile(0.75)), 2) if rent_observations else None,
                "rent_measure": "observed_bond_lodgement_weekly_rent",
                "currency": "AUD",
                "source_dataset": SOURCE_DATASET,
                "source_url": DATA_PAGE_URL,
            }
        )
    return pd.DataFrame(records)


def attach_ghsl_match(panel: pd.DataFrame, city_spine_path: Path) -> pd.DataFrame:
    if not city_spine_path.exists():
        raise FileNotFoundError(f"missing city spine: {city_spine_path}")
    spine = (
        pd.read_csv(city_spine_path)
        if city_spine_path.suffix.lower() == ".csv"
        else pd.read_parquet(city_spine_path)
    )
    au = spine[spine["country_name"].eq("Australia")].copy()
    name_to_match: dict[str, dict[str, Any]] = {}
    for row in au.to_dict("records"):
        name_to_match[str(row["city_name"]).strip().lower()] = {
            "ieset_city_id": row["ieset_city_id"],
            "ghsl_city_name": row["city_name"],
            "ghsl_city_rank_2025": int(row["city_rank_2025"]),
        }

    out_rows = []
    for geography_id in panel["geography_id"].drop_duplicates():
        ghsl_city = GEOGRAPHY_GHSL_CITY_NAME.get(geography_id)
        match = name_to_match.get(ghsl_city.strip().lower()) if ghsl_city else None
        if match:
            out_rows.append(
                {
                    "geography_id": geography_id,
                    "ieset_city_id": match["ieset_city_id"],
                    "ghsl_city_name": match["ghsl_city_name"],
                    "ghsl_city_rank_2025": match["ghsl_city_rank_2025"],
                    "ghsl_match_flag": True,
                    "ghsl_match_type": "manual_metro_aggregate_to_ghsl_urban_centre",
                    "manual_review_required": False,
                }
            )
        else:
            out_rows.append(
                {
                    "geography_id": geography_id,
                    "ieset_city_id": None,
                    "ghsl_city_name": None,
                    "ghsl_city_rank_2025": pd.NA,
                    "ghsl_match_flag": False,
                    "ghsl_match_type": "no_ghsl_urban_centre_aggregate",
                    "manual_review_required": False,
                }
            )
    return panel.merge(pd.DataFrame(out_rows), on="geography_id", how="left")


def build_panel(
    *,
    city_spine_path: Path,
    lodgement_urls: list[str] | None = None,
    input_files: list[Path] | None = None,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    source_shas: dict[str, str] = {}
    micro_frames: list[pd.DataFrame] = []

    if input_files:
        for path in input_files:
            payload = path.read_bytes()
            source_shas[path.name] = sha256_bytes(payload)
            micro_frames.append(read_lodgement_workbook(payload, source_url=DATA_PAGE_URL, source_file=path.name))
    else:
        urls = lodgement_urls if lodgement_urls is not None else DEFAULT_LODGEMENT_URLS
        for url in urls:
            payload = fetch_bytes(url)
            name = url.rsplit("/", 1)[-1]
            source_shas[name] = sha256_bytes(payload)
            micro_frames.append(read_lodgement_workbook(payload, source_url=url, source_file=name))

    micro = pd.concat(micro_frames, ignore_index=True)
    if micro.empty:
        raise ValueError("NSW rental bond workbooks yielded no lodgement rows")

    panel = aggregate_panel(micro)
    panel = attach_ghsl_match(panel, city_spine_path)

    ordered = [
        "jurisdiction_code",
        "jurisdiction_name",
        "country_name",
        "country_iso3",
        "period",
        "period_type",
        "geography_id",
        "geography_label",
        "ieset_city_id",
        "ghsl_city_name",
        "ghsl_city_rank_2025",
        "ghsl_match_flag",
        "ghsl_match_type",
        "manual_review_required",
        "dwelling_type_code",
        "dwelling_type_label",
        "bedroom_group",
        "bond_lodgement_count",
        "rent_observation_count",
        "median_weekly_rent_aud",
        "mean_weekly_rent_aud",
        "p25_weekly_rent_aud",
        "p75_weekly_rent_aud",
        "rent_measure",
        "currency",
        "source_dataset",
        "source_url",
    ]
    panel = (
        panel[ordered]
        .sort_values(["period", "geography_id", "dwelling_type_code", "bedroom_group"])
        .reset_index(drop=True)
    )

    matched = panel["ghsl_match_flag"].fillna(False)
    stats = {
        "panel_rows": int(len(panel)),
        "micro_lodgement_rows": int(len(micro)),
        "start_period": str(panel["period"].min()),
        "end_period": str(panel["period"].max()),
        "period_count": int(panel["period"].nunique()),
        "geography_count": int(panel["geography_id"].nunique()),
        "matched_geographies": int(panel.loc[matched, "geography_id"].nunique()),
        "matched_panel_rows": int(matched.sum()),
        "unique_ieset_city_ids": int(panel["ieset_city_id"].nunique(dropna=True)),
        "source_file_sha256": source_shas,
        "rent_measure": "observed_bond_lodgement_weekly_rent",
        "geography_method": (
            "Greater Sydney derived from ABS GCCSA postcode ranges applied to NSW Fair "
            "Trading lodgement micro-data; nsw_state covers all lodgements."
        ),
    }
    return panel, stats


def manifest_entry(result: FetchResult) -> dict[str, Any]:
    payload = asdict(result)
    payload["fetch_utc"] = result.fetch_utc.isoformat()
    payload["parquet_path"] = rel(result.parquet_path)
    return payload


def write_manifest(result: FetchResult, manifest_dir: Path, run_stamp: str) -> Path:
    manifest_dir.mkdir(parents=True, exist_ok=True)
    path = manifest_dir / f"fetch_run_{run_stamp}_australia_rental_bond.yaml"
    payload = {"run_utc": run_stamp, "pipeline": "australia_rental_bond_panel", "entries": [manifest_entry(result)]}
    path.write_text(yaml.safe_dump(payload, sort_keys=False, allow_unicode=True))
    return path


def emit(panel: pd.DataFrame, stats: dict[str, Any], output_path: Path, fetch_ts: datetime) -> FetchResult:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    panel.to_parquet(output_path, engine="pyarrow", index=False)
    parquet_sha = sha256_path(output_path)
    return FetchResult(
        publisher=PUBLISHER,
        series_id=SERIES_ID,
        source_url=DATA_PAGE_URL,
        methodology_url=METHODOLOGY_URL,
        license=LICENSE,
        fetch_utc=fetch_ts,
        rows=len(panel),
        frequency="monthly lodgement batches",
        units="observed weekly bond-lodgement rents (AUD) and bond lodgement counts",
        currency="AUD",
        start_date=stats["start_period"],
        end_date=stats["end_period"],
        sha256=parquet_sha,
        parquet_path=output_path,
        extra={
            "artifacts": {"parquet_path": rel(output_path), "parquet_sha256": parquet_sha},
            "stats": stats,
            "columns": list(panel.columns),
            "construction": (
                "NSW Fair Trading monthly rental-bond lodgement XLSX micro-data (one bond per row: "
                "lodgement date, postcode, dwelling type, bedrooms, weekly rent) aggregated into median "
                "weekly observed bond-lodgement rents and lodgement counts by (jurisdiction, period, "
                "geography, dwelling type, bedroom group). Greater Sydney is a GCCSA postcode-derived "
                "aggregate crosswalked to the Sydney GHSL urban-centre id."
            ),
        },
    )


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--city-spine", default="data/derived/city_universe_top1000.parquet")
    parser.add_argument("--input", nargs="*", help="Optional local NSW lodgement XLSX files; otherwise fetch defaults.")
    parser.add_argument("--url", nargs="*", help="Optional explicit lodgement XLSX URLs to fetch.")
    parser.add_argument("--output", default="data/derived/australia_rental_bond_panel.parquet")
    parser.add_argument("--manifest-dir", default="data/manifests")
    parser.add_argument("--fetch-utc", help="Optional UTC timestamp for deterministic builds/tests.")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    fetch_ts = parse_fetch_utc(args.fetch_utc)
    input_files = [path_arg(p).resolve() for p in args.input] if args.input else None
    panel, stats = build_panel(
        city_spine_path=path_arg(args.city_spine).resolve(),
        lodgement_urls=args.url,
        input_files=input_files,
    )
    output_path = path_arg(args.output).resolve()
    result = emit(panel, stats, output_path, fetch_ts)
    manifest = write_manifest(result, path_arg(args.manifest_dir).resolve(), utc_stamp(fetch_ts))
    print(
        f"OK {PUBLISHER}:{SERIES_ID} rows={result.rows} "
        f"period={result.start_date}->{result.end_date} matched_cities={stats['unique_ieset_city_ids']}"
    )
    print(f"artifact: {rel(output_path)} sha256={result.sha256}")
    print(f"manifest: {rel(manifest)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
